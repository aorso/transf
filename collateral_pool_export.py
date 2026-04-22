# -*- coding: utf-8 -*-
"""
Export Collatéral Pool – génère un fichier Excel formaté à partir du rapport TXT.

Nom du fichier de sortie : YYYYMMDD - Collatéral Pool.xlsx
  (la date est extraite du nom du fichier source au format …DDMMYYYY)
"""

import csv
import re
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ─── 1. CHARGEMENT DU FICHIER TXT ────────────────────────────────────────────

def load_302_303(path) -> pd.DataFrame:
    """Charge les lignes 302 (header) et 303 (données) du fichier positionnel."""
    path = Path(path)
    header = None
    data = []

    with path.open("r", encoding="utf-8", newline="") as f:
        for raw in f:
            line = raw.rstrip("\r\n")
            if not line:
                continue
            if not (line.startswith('"302",') or line.startswith('"303",')):
                continue
            fields = next(csv.reader([line], delimiter=",", quotechar='"'))
            code = fields[0].strip('"')
            if code == "302":
                header = fields
            elif code == "303":
                data.append(fields)

    if header is None:
        raise ValueError("Pas de ligne 302 (header) dans le fichier.")
    if not data:
        raise ValueError("Pas de ligne 303 (données) dans le fichier.")

    return pd.DataFrame(data, columns=header)


# ─── 2. EXTRACTION DE LA DATE DEPUIS LE NOM DU FICHIER ───────────────────────

def extract_date_yyyymmdd(path) -> str:
    """
    Extrait la date au format DDMMYYYY en fin de nom de fichier
    et la retourne au format YYYYMMDD.

    Ex. : "FULL POSITIONS REPORT_Confidential 22042026.TXT"
          → "20260422"
    """
    stem = Path(path).stem
    match = re.search(r'(\d{2})(\d{2})(\d{4})\s*$', stem)
    if not match:
        raise ValueError(
            f"Impossible d'extraire une date DDMMYYYY depuis le nom de fichier : '{stem}'"
        )
    dd, mm, yyyy = match.groups()
    return f"{yyyy}{mm}{dd}"


# ─── 3. CONSTANTES DE STYLE ──────────────────────────────────────────────────

_GREEN_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
_WHITE_BOLD  = Font(bold=True, color="FFFFFF")
_THIN_BLACK  = Side(style="thin", color="000000")
_BLACK_BORDER = Border(
    left=_THIN_BLACK, right=_THIN_BLACK,
    top=_THIN_BLACK,  bottom=_THIN_BLACK,
)


# ─── 4. FORMATAGE D'UNE FEUILLE ──────────────────────────────────────────────

def format_sheet(ws, df: pd.DataFrame, pct_cols: list | None = None):
    """
    Applique le formatage à une feuille openpyxl :
      - Quadrillage noir uniquement sur les cellules du tableau
      - En-tête : fond vert, police blanche/gras, centré
      - Colonnes ajustées au contenu
      - Colonnes listées dans pct_cols : format pourcentage (0,00 %)
      - Quadrillage Excel (fond de page) désactivé
    """
    n_rows, n_cols = df.shape

    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        col_name   = df.columns[col_idx - 1]

        # ── En-tête (ligne 1) ────────────────────────────────────────────────
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.font      = _WHITE_BOLD
        header_cell.fill      = _GREEN_FILL
        header_cell.border    = _BLACK_BORDER
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Données (lignes 2 … n+1) ─────────────────────────────────────────
        is_pct = pct_cols and col_name in pct_cols
        for row_idx in range(2, n_rows + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BLACK_BORDER
            if is_pct:
                cell.number_format = "0.00%"

        # ── Largeur de colonne : max(header, valeurs) + marge ────────────────
        all_values = [str(col_name)] + [
            str(v) for v in df.iloc[:, col_idx - 1]
        ]
        max_len = max(len(v) for v in all_values)
        ws.column_dimensions[col_letter].width = max_len + 4

    # Désactiver le quadrillage de fond (hors tableau, la page reste blanche)
    ws.sheet_view.showGridLines = False


# ─── 5. PIPELINE PRINCIPAL ───────────────────────────────────────────────────

file_path = Path("FULL POSITIONS REPORT_Confidential.TXT")   # ← adapter si besoin

# Chargement
df = load_302_303(file_path)

# Filtrage sur le compte collatéral 25570
df = df[df["Collateral Account"] == "25570"].copy()
df = df[["ISIN Code", "Security Name", "Marginal Value"]]

# Nettoyage et calcul des poids
df["Marginal Value"] = pd.to_numeric(
    df["Marginal Value"].str.replace(",", "."), errors="coerce"
)
total         = df["Marginal Value"].sum()
df["Weight"]  = df["Marginal Value"] / total

# ── Tableau ALL : liste complète (Security Name + ISIN) ──────────────────────
tab0 = df[["Security Name", "ISIN Code"]].copy()
tab0.rename(columns={"ISIN Code": "ISIN"}, inplace=True)

# ── Tableau Table : Top 20 par poids ─────────────────────────────────────────
df_top20 = df.nlargest(20, "Weight")
tab1 = df_top20[["Security Name", "ISIN Code"]].copy()
tab1.rename(
    columns={"Security Name": "Top 20 Securities", "ISIN Code": "ISIN"},
    inplace=True,
)

# ── Tableau % Diversification : Rank + % Poids des Top 20 ────────────────────
tab2 = df_top20[["Weight"]].copy()
tab2.rename(columns={"Weight": "% Diversification"}, inplace=True)
tab2.insert(0, "Rank", range(1, len(tab2) + 1))

# ── Nom du fichier de sortie ──────────────────────────────────────────────────
date_str    = extract_date_yyyymmdd(file_path)
output_path = Path(file_path.parent, f"{date_str} - Collatéral Pool.xlsx")

# ── Export Excel + formatage ──────────────────────────────────────────────────
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    tab0.to_excel(writer, sheet_name="ALL",               index=False)
    tab1.to_excel(writer, sheet_name="Table",             index=False)
    tab2.to_excel(writer, sheet_name="% Diversification", index=False)

    format_sheet(writer.sheets["ALL"],               tab0)
    format_sheet(writer.sheets["Table"],             tab1)
    format_sheet(
        writer.sheets["% Diversification"],
        tab2,
        pct_cols=["% Diversification"],
    )

print(f"✓ Fichier généré : {output_path}")
