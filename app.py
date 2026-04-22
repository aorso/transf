# -*- coding: utf-8 -*-
"""
Streamlit – Collatéral Pool Exporter
Glisser-déposer le fichier TXT → télécharger le fichier Excel formaté.
"""

import csv
import io
import re
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ─── STYLES EXCEL ────────────────────────────────────────────────────────────

_GREEN_FILL   = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
_WHITE_BOLD   = Font(bold=True, color="FFFFFF")
_THIN_BLACK   = Side(style="thin", color="000000")
_BLACK_BORDER = Border(
    left=_THIN_BLACK, right=_THIN_BLACK,
    top=_THIN_BLACK,  bottom=_THIN_BLACK,
)


# ─── FONCTIONS MÉTIER ────────────────────────────────────────────────────────

def load_302_303(file_bytes: bytes) -> pd.DataFrame:
    """Parse les lignes 302 (header) et 303 (données) depuis les bytes du fichier."""
    header, data = None, []
    text = file_bytes.decode("utf-8", errors="replace")

    for raw in text.splitlines():
        line = raw.rstrip("\r\n")
        if not line:
            continue
        if not (line.startswith('"302",') or line.startswith('"303",')):
            continue
        fields = next(csv.reader([line], delimiter=",", quotechar='"'))
        code = fields[0].strip('"')
        if code == "302":
            header = fields
        elif code == "303":
            data.append(fields)

    if header is None:
        raise ValueError("Pas de ligne 302 (header) dans le fichier.")
    if not data:
        raise ValueError("Pas de ligne 303 (données) dans le fichier.")

    return pd.DataFrame(data, columns=header)


def extract_date_yyyymmdd(filename: str) -> str:
    """
    Extrait DDMMYYYY en fin de nom de fichier → retourne YYYYMMDD.
    Ex. : "FULL POSITIONS REPORT_Confidential 22042026.TXT" → "20260422"
    """
    stem = filename.rsplit(".", 1)[0]
    match = re.search(r'(\d{2})(\d{2})(\d{4})\s*$', stem)
    if not match:
        raise ValueError(
            f"Impossible d'extraire une date DDMMYYYY depuis le nom : '{stem}'"
        )
    dd, mm, yyyy = match.groups()
    return f"{yyyy}{mm}{dd}"


def format_sheet(ws, df: pd.DataFrame, pct_cols: list | None = None):
    """Formate une feuille openpyxl : en-tête vert, bordures noires, colonnes ajustées."""
    n_rows, n_cols = df.shape

    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        col_name   = df.columns[col_idx - 1]

        # En-tête
        hc = ws.cell(row=1, column=col_idx)
        hc.font      = _WHITE_BOLD
        hc.fill      = _GREEN_FILL
        hc.border    = _BLACK_BORDER
        hc.alignment = Alignment(horizontal="center", vertical="center")

        # Données
        is_pct = pct_cols and col_name in pct_cols
        for row_idx in range(2, n_rows + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BLACK_BORDER
            if is_pct:
                cell.number_format = "0.00%"

        # Largeur auto
        all_values = [str(col_name)] + [str(v) for v in df.iloc[:, col_idx - 1]]
        ws.column_dimensions[col_letter].width = max(len(v) for v in all_values) + 4

    ws.sheet_view.showGridLines = False


def build_excel(file_bytes: bytes, filename: str) -> tuple[bytes, str]:
    """
    Exécute tout le pipeline et retourne (excel_bytes, output_filename).
    """
    # Chargement et filtrage
    df = load_302_303(file_bytes)
    df = df[df["Collateral Account"] == "25570"].copy()
    df = df[["ISIN Code", "Security Name", "Marginal Value"]]
    df["Marginal Value"] = pd.to_numeric(
        df["Marginal Value"].str.replace(",", "."), errors="coerce"
    )
    total        = df["Marginal Value"].sum()
    df["Weight"] = df["Marginal Value"] / total

    # Tableaux
    tab0 = df[["Security Name", "ISIN Code"]].copy()
    tab0.rename(columns={"ISIN Code": "ISIN"}, inplace=True)

    df_top20 = df.nlargest(20, "Weight")
    tab1 = df_top20[["Security Name", "ISIN Code"]].copy()
    tab1.rename(columns={"Security Name": "Top 20 Securities", "ISIN Code": "ISIN"}, inplace=True)

    tab2 = df_top20[["Weight"]].copy()
    tab2.rename(columns={"Weight": "% Diversification"}, inplace=True)
    tab2.insert(0, "Rank", range(1, len(tab2) + 1))

    # Nom du fichier de sortie
    date_str      = extract_date_yyyymmdd(filename)
    output_name   = f"{date_str} - Collatéral Pool.xlsx"

    # Génération Excel en mémoire
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        tab0.to_excel(writer, sheet_name="ALL",               index=False)
        tab1.to_excel(writer, sheet_name="Table",             index=False)
        tab2.to_excel(writer, sheet_name="% Diversification", index=False)

        format_sheet(writer.sheets["ALL"],               tab0)
        format_sheet(writer.sheets["Table"],             tab1)
        format_sheet(
            writer.sheets["% Diversification"],
            tab2,
            pct_cols=["% Diversification"],
        )

    return buffer.getvalue(), output_name


# ─── INTERFACE STREAMLIT ─────────────────────────────────────────────────────

st.set_page_config(
    page_title="Collatéral Pool Exporter",
    page_icon="📊",
    layout="centered",
)

st.title("📊 Collatéral Pool Exporter")
st.markdown(
    "Déposez votre fichier **FULL POSITIONS REPORT** (`.TXT`) ci-dessous "
    "pour générer l'Excel formaté."
)

uploaded = st.file_uploader(
    label="Glisser-déposer ou sélectionner le fichier TXT",
    type=["txt", "TXT"],
    help="Le nom du fichier doit se terminer par une date au format DDMMYYYY",
)

if uploaded is not None:
    st.info(f"Fichier reçu : **{uploaded.name}**")

    with st.spinner("Traitement en cours…"):
        try:
            excel_bytes, output_name = build_excel(uploaded.read(), uploaded.name)

            st.success(f"Fichier prêt : **{output_name}**")

            # Aperçu des données
            with st.expander("Aperçu des tableaux générés"):
                df_preview = load_302_303(uploaded.getvalue())
                df_preview = df_preview[df_preview["Collateral Account"] == "25570"].copy()
                df_preview = df_preview[["ISIN Code", "Security Name", "Marginal Value"]]
                df_preview["Marginal Value"] = pd.to_numeric(
                    df_preview["Marginal Value"].str.replace(",", "."), errors="coerce"
                )
                total = df_preview["Marginal Value"].sum()
                df_preview["Weight"] = df_preview["Marginal Value"] / total

                tab_all, tab_top20, tab_pct = st.tabs(["ALL", "Table (Top 20)", "% Diversification"])
                with tab_all:
                    st.dataframe(
                        df_preview[["Security Name", "ISIN Code"]].rename(columns={"ISIN Code": "ISIN"}),
                        use_container_width=True,
                    )
                with tab_top20:
                    top20 = df_preview.nlargest(20, "Weight")
                    st.dataframe(
                        top20[["Security Name", "ISIN Code"]].rename(
                            columns={"Security Name": "Top 20 Securities", "ISIN Code": "ISIN"}
                        ),
                        use_container_width=True,
                    )
                with tab_pct:
                    pct = df_preview.nlargest(20, "Weight")[["Weight"]].copy()
                    pct.rename(columns={"Weight": "% Diversification"}, inplace=True)
                    pct.insert(0, "Rank", range(1, len(pct) + 1))
                    pct["% Diversification"] = pct["% Diversification"].map("{:.2%}".format)
                    st.dataframe(pct, use_container_width=True)

            st.download_button(
                label="⬇️  Télécharger l'Excel",
                data=excel_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"Erreur lors du traitement : {e}")
